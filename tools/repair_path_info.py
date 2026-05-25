#!/usr/bin/env python
"""
M8 路径修正工具（V2）— 处理路径信息 Excel 中的原始路径

读取 research/analysis/路径信息.xlsx，使用 M8 V2 算法（锚点 DP + 分段 KSP）
进行拓扑合法性检查和补全，输出修正结果到 Excel。
若 V2 不可行则自动回退 V1（相邻补全）。

支持生成可视化图片，展示原始路径与修正路径的对比。

用法：
    # 默认处理原路径列（去重后唯一路径，V2 算法）
    uv run python tools/repair_path_info.py

    # 处理全部 415 条
    uv run python tools/repair_path_info.py --all

    # 限制条数（调试用）
    uv run python tools/repair_path_info.py --limit 50

    # 自定义输出路径
    uv run python tools/repair_path_info.py --output outputs/路径修正结果.xlsx

    # 跳过拓扑加载（仅做格式清洗，不做最短路径补全）
    uv run python tools/repair_path_info.py --skip-topology

    # 生成可视化图片（默认保存到 outputs/path_viz/）
    uv run python tools/repair_path_info.py --visualize

    # 自定义可视化输出目录
    uv run python tools/repair_path_info.py --visualize --viz-output outputs/my_viz/

    # 可视化前 20 条
    uv run python tools/repair_path_info.py --visualize --viz-limit 20

    # 关闭 V2 算法（仅用 V1 相邻补全）
    # 需修改脚本中 config["use_v2_algorithm"] = False
"""

import argparse
import sys
import time
from pathlib import Path
from typing import Optional

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from dotenv import load_dotenv

load_dotenv(project_root / ".env")

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.modules.m8_path_repair.core.graph import RoadGraph
from src.modules.m8_path_repair.core.pipeline import repair_single
from src.modules.m8_path_repair.core.normalizer import normalize_raw_path
from src.modules.m8_path_repair.core.metrics import (
    calc_repeated_node_count,
    calc_reverse_edge_count,
    calc_detour_ratio,
    calc_backtrack_index,
    calc_repair_confidence,
    get_repair_status,
    calc_path_length_meters,
)
from src.modules.m8_path_repair.core.visualizer import (
    visualize_repair_result,
    visualize_comparison,
)

from src.app.logger import get_logger, setup_logging

logger = get_logger(__name__)

INPUT_FILE = project_root / "research" / "analysis" / "路径信息.xlsx"

# 输出样式常量
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_FILL = {
    "HIGH_CONFIDENCE": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "MEDIUM_CONFIDENCE": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "LOW_CONFIDENCE": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "NEED_MANUAL_REVIEW": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "FAILED_NO_PATH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "FAILED_EMPTY_PATH": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="M8 路径修正工具 — 处理路径信息 Excel")
    parser.add_argument("--all", action="store_true", help="处理全部数据（默认只处理去重后的唯一路径）")
    parser.add_argument("--limit", type=int, default=None, help="限制处理条数")
    parser.add_argument("--output", type=str, default=None, help="输出 Excel 文件路径")
    parser.add_argument("--skip-topology", action="store_true", help="跳过拓扑加载，仅做格式清洗")
    parser.add_argument("--topology-version", default="202512", help="拓扑数据版本")
    # 可视化参数
    parser.add_argument("--visualize", action="store_true", help="生成可视化图片")
    parser.add_argument("--viz-output", type=str, default=None, help="可视化输出目录")
    parser.add_argument("--viz-limit", type=int, default=None, help="可视化图片数量限制")
    # 往复路径过滤参数
    parser.add_argument("--remove-reciprocal", action="store_true", help="去除往复路径片段（A->B->A形式）")
    # 施工路段参数
    parser.add_argument("--construction-sections", type=str, default="G007061002000520", help="施工收费单元列表（用|分隔）")
    # 车型参数
    parser.add_argument("--vehicle-type", type=int, default=11, help="车型（默认11=货车）")
    return parser.parse_args()


def read_input_excel(path: Path) -> tuple[list[str], list[dict]]:
    """
    读取输入 Excel。

    Returns:
        (headers, records)
        headers: 列名列表
        records: 每行数据的字典列表
    """
    logger.info(f"读取输入文件: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("Excel 文件为空")

    headers = [str(h) for h in rows[0]]
    records = []
    for row in rows[1:]:
        rec = {}
        for i, h in enumerate(headers):
            rec[h] = row[i] if i < len(row) else None
        records.append(rec)

    logger.info(f"读取到 {len(records)} 行数据，列: {headers}")
    return headers, records


def run_repair(
    records: list[dict],
    graph: Optional[RoadGraph],
    config: dict,
) -> list[dict]:
    """
    对每条记录的原路径进行修正。

    Returns:
        增强后的记录列表，每行增加了修正结果字段
    """
    results = []
    skipped = 0

    for idx, rec in enumerate(records):
        raw_path = rec.get("原路径", "")
        if not raw_path or str(raw_path).strip() == "":
            skipped += 1
            results.append({
                **rec,
                "_status": "SKIPPED",
                "_reason": "原路径为空",
            })
            continue

        raw_path = str(raw_path).strip()
        enid = rec.get("OD起点门架", "") or ""
        exid = rec.get("OD终点门架", "") or ""
        record_id = f"row_{idx + 1}"

        if graph is None:
            # 跳过拓扑模式：仅做标准化
            from src.modules.m8_path_repair.core.normalizer import (
                split_raw_path,
                remove_consecutive_duplicates,
            )
            nodes = split_raw_path(raw_path)
            deduped, dup_count = remove_consecutive_duplicates(nodes)
            corrected = "|".join(deduped)
            results.append({
                **rec,
                "_corrected_path": corrected,
                "_status": "FORMAT_CLEAN_ONLY",
                "_raw_node_count": len(nodes),
                "_corrected_node_count": len(deduped),
                "_consecutive_dups": dup_count,
                "_repair_confidence": 50.0,
                "_repair_status": "FORMAT_CLEAN_ONLY",
            })
            continue

        try:
            result = repair_single(
                record_id=record_id,
                enid=str(enid) if enid else "",
                exid=str(exid) if exid else "",
                raw_path=raw_path,
                graph=graph,
                config=config,
            )
            results.append({
                **rec,
                "record_id": record_id,  # 添加 record_id 用于可视化
                "_corrected_path": result["corrected_path"],
                "_status": "REPAIRED",
                "_raw_node_count": result["raw_node_count"],
                "_corrected_node_count": result["corrected_node_count"],
                "_inserted_node_count": result["inserted_node_count"],
                "_dropped_node_count": result["dropped_node_count"],
                "_raw_match_ratio": result["raw_match_ratio"],
                "_detour_ratio": result["detour_ratio"],
                "_backtrack_index": result["backtrack_index"],
                "_repair_confidence": result["repair_confidence"],
                "_repair_status": result["repair_status"],
                "_repair_detail": str(result.get("repair_detail", {})),
                # 最短路径与费额相关字段
                "shortest_path": result.get("shortest_path", ""),
                "shortest_path_distance": result.get("shortest_path_distance"),
                "min_fee_path": result.get("min_fee_path", ""),
                "min_fee": result.get("min_fee"),
                "passes_construction": result.get("passes_construction"),
                "shortest_through_construction": result.get("shortest_through_construction"),
                # 绕行路径（绕开施工路段）
                "detour_path": result.get("detour_path", ""),
                "detour_distance": result.get("detour_distance"),
                "detour_fee": result.get("detour_fee"),
            })
        except Exception as e:
            logger.warning(f"Row {idx + 1} 修正失败: {e}")
            results.append({
                **rec,
                "_status": "ERROR",
                "_reason": str(e),
            })

    if skipped > 0:
        logger.info(f"跳过 {skipped} 条空路径")

    return results


def write_output_excel(
    path: Path,
    original_headers: list[str],
    results: list[dict],
) -> None:
    """
    写入输出 Excel，带样式。
    """
    logger.info(f"写入输出文件: {path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "路径修正结果"

    # 输出列：原始列 + 修正结果列
    output_headers = original_headers + [
        "修正后路径",
        "原始节点数",
        "修正后节点数",
        "插入节点数",
        "删除节点数",
        "原始匹配率",
        "绕行比",
        "折返指数",
        "修正置信度",
        "修正状态",
        "修正详情",
        "处理状态",
        # 最短路径与费额相关字段
        "最短路径",
        "最短路径距离(米)",
        "最小费额路径",
        "最小费额(元)",
        "是否经过施工路段",
        "施工路段最短距离(米)",
        # 绕行路径（绕开施工路段）
        "绕行路径(不经过施工)",
        "绕行距离(米)",
        "绕行费额(元)",
    ]

    col_count = len(output_headers)

    # 写表头
    for col_idx, header in enumerate(output_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 写数据
    for row_idx, rec in enumerate(results, 2):
        for col_idx, header in enumerate(output_headers, 1):
            # 映射列名
            if header in original_headers:
                value = rec.get(header, "")
            else:
                # 修正结果列
                field_map = {
                    "修正后路径": "_corrected_path",
                    "原始节点数": "_raw_node_count",
                    "修正后节点数": "_corrected_node_count",
                    "插入节点数": "_inserted_node_count",
                    "删除节点数": "_dropped_node_count",
                    "原始匹配率": "_raw_match_ratio",
                    "绕行比": "_detour_ratio",
                    "折返指数": "_backtrack_index",
                    "修正置信度": "_repair_confidence",
                    "修正状态": "_repair_status",
                    "修正详情": "_repair_detail",
                    "处理状态": "_status",
                    # 最短路径与费额相关字段（直接从结果字典读取，无前缀）
                    "最短路径": "shortest_path",
                    "最短路径距离(米)": "shortest_path_distance",
                    "最小费额路径": "min_fee_path",
                    "最小费额(元)": "min_fee",
                    "是否经过施工路段": "passes_construction",
                    "施工路段最短距离(米)": "shortest_through_construction",
                    # 绕行路径（绕开施工路段）
                    "绕行路径(不经过施工)": "detour_path",
                    "绕行距离(米)": "detour_distance",
                    "绕行费额(元)": "detour_fee",
                }
                field = field_map.get(header)
                value = rec.get(field, "") if field else ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

            # 修正状态列着
            if header == "修正状态" and value in STATUS_FILL:
                cell.fill = STATUS_FILL[value]

    # 设置列宽
    col_widths = {}
    for header in output_headers:
        if header == "原路径" or header == "修正后路径":
            col_widths[header] = 60
        elif header == "修正详情":
            col_widths[header] = 40
        elif header in ("OD起点门架", "OD终点门架"):
            col_widths[header] = 22
        else:
            col_widths[header] = 15

    for col_idx, header in enumerate(output_headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    logger.info(f"输出文件已保存: {path}")


def main() -> int:
    setup_logging()
    args = parse_args()

    logger.info("=" * 60)
    logger.info("M8 路径修正工具 — 路径信息 Excel 处理")
    logger.info("=" * 60)

    start_time = time.time()

    # 读取输入
    headers, records = read_input_excel(INPUT_FILE)

    # 去重（默认只处理唯一路径）
    if not args.all:
        seen_paths = set()
        unique_records = []
        for rec in records:
            raw = str(rec.get("原路径", "")).strip()
            if raw and raw not in seen_paths:
                seen_paths.add(raw)
                unique_records.append(rec)
        logger.info(f"去重前 {len(records)} 行，去重后 {len(unique_records)} 条唯一路径")
        records = unique_records

    if args.limit:
        records = records[: args.limit]
        logger.info(f"限制处理: {len(records)} 条")

    # 加载图
    graph = None
    if not args.skip_topology:
        logger.info("加载拓扑图...")
        graph = RoadGraph(version=args.topology_version)
        graph.load_topology_cache()
        graph.load_node_info()
        logger.info(f"拓扑图加载完成: {len(graph._next_cache)} 条边, {len(graph._node_info)} 个节点")

    config = {
        # V2 算法开关
        "use_v2_algorithm": True,

        # V1 参数
        "max_gap_search_window": 6,
        "max_detour_ratio": 2.0,
        "backward_progress_threshold_m": 300.0,

        # V2 锚点筛选
        "max_anchor_detour_ratio": 1.5,
        "skip_node_penalty": 50.0,

        # V2 分段 KSP
        "k_shortest_paths": 5,
        "segment_score_weights": {
            "path_cost": 0.40,
            "observation_match": 0.30,
            "inserted_nodes": 0.15,
            "backtrack": 0.10,
            "skipped_nodes": 0.05,
        },

        # V2 后处理
        "allow_uturn_count": 1,

        # 往复路径过滤
        "remove_reciprocal_path": args.remove_reciprocal,

        # 施工路段
        "construction_sections": args.construction_sections,

        # 车型
        "vehicle_type": args.vehicle_type,

        # 输出控制
        "detail_geo": False,  # Excel 输出不需要经纬度
    }

    # 执行修正
    results = run_repair(records, graph, config)

    # 统计
    status_counts = {}
    for r in results:
        s = r.get("_status", "UNKNOWN")
        status_counts[s] = status_counts.get(s, 0) + 1

    logger.info("处理完成统计:")
    for status, count in sorted(status_counts.items()):
        logger.info(f"  {status}: {count}")

    # 输出
    output_path = Path(args.output) if args.output else project_root / "outputs" / "路径修正结果.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_output_excel(output_path, headers, results)

    elapsed = time.time() - start_time
    logger.info(f"总耗时: {elapsed:.2f}s")

    # 生成可视化图片
    if args.visualize and graph is not None:
        logger.info("=" * 60)
        logger.info("生成可视化图片...")
        viz_start = time.time()

        # 筛选成功修正的结果
        viz_results = [
            r for r in results
            if r.get("_status") == "REPAIRED" and r.get("_corrected_path")
        ]

        if not viz_results:
            logger.warning("没有可可视化的修正结果")
        else:
            viz_output_dir = Path(args.viz_output) if args.viz_output else project_root / "outputs" / "path_viz"
            viz_output_dir.mkdir(parents=True, exist_ok=True)

            # 只对成功修正的结果生成可视化
            viz_count = args.viz_limit if args.viz_limit else len(viz_results)
            viz_count = min(viz_count, len(viz_results))

            # 将结果字典转换为可视化所需的格式
            viz_input = []
            for r in viz_results[:viz_count]:
                # 构建与 repair_single 输出格式一致的结果
                viz_input.append({
                    "record_id": r.get("record_id", "unknown"),
                    "enid": r.get("OD起点门架", "") or "",
                    "exid": r.get("OD终点门架", "") or "",
                    "raw_path": r.get("原路径", ""),
                    "corrected_path": r.get("_corrected_path", ""),
                    "raw_node_count": r.get("_raw_node_count", 0),
                    "corrected_node_count": r.get("_corrected_node_count", 0),
                    "inserted_node_count": r.get("_inserted_node_count", 0),
                    "dropped_node_count": r.get("_dropped_node_count", 0),
                    "raw_match_ratio": r.get("_raw_match_ratio", 0),
                    "detour_ratio": r.get("_detour_ratio", 0),
                    "backtrack_index": r.get("_backtrack_index", 0),
                    "repair_confidence": r.get("_repair_confidence", 0),
                    "repair_status": r.get("_repair_status", "UNKNOWN"),
                    "reverse_edge_count": 0,
                    "backward_progress_count": 0,
                    "u_turn_count": 0,
                    "repeated_node_count": 0,
                })

            output_paths = visualize_comparison(
                results=viz_input,
                graph=graph,
                output_dir=viz_output_dir,
                prefix="path_repair",
            )

            logger.info(f"已生成 {len(output_paths)} 张可视化图片")
            logger.info(f"可视化保存目录: {viz_output_dir}")

        viz_elapsed = time.time() - viz_start
        logger.info(f"可视化耗时: {viz_elapsed:.2f}s")

    logger.info("=" * 60)

    if graph:
        graph.close()

    return 0


if __name__ == "__main__":
    sys.exit(main())
